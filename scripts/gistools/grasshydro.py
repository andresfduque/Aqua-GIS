#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Created by AndresD at 8/06/19.

Module that integrates best tools for geomorphometric analysis available in different GIS open source software

Features:
    + Curve number estimation
    + Raster accumulation weighting
    + Extract raster value
    + Voronoi interpolation
    + Basin polygon weighting

Pre-requisites:
    + GRASS python scripting library

@author:    Andres Felipe Duque Perez
Email:      aduquep@ingenevo.com.co

Credits:

GRASS Development Team, 2019. Geographic Resources Analysis Support System (GRASS) Software, Version 7.8.
Open Source Geospatial Foundation. https://grass.osgeo.org
"""
# modules import
import sqlite3
from time import time

import numpy as np
import pandas as pd
from openpyxl import load_workbook

import grass.script as grass
import grass.script.array as garray


def r_curve_number_estimation(cover_shapefile, clc_field_name, soil_field_name, lookup_file_path, curve_number_r):
    """Estimate NRCS Runoff Curve Number.

    Parameters
    ----------
    cover_shapefile: str
        Shapefile containing land cover and soil parameters
    clc_field_name: str
        Name of field containing Corine Land Cover code
    soil_field_name: str
        Name of field containing Soil type code
    lookup_file_path: str
        File containing relationship between land cover, soil type and curve number [excel predefined format]
    curve_number_r: str
        Curve number map

    Returns
    -------
    Add Curve Number column in cover shapefile
    Curve Number map

    """
    # ==================================================================================================================
    # -- Estimate CN
    # ==================================================================================================================
    start_time = time()

    # read shapefile table
    sql_path = grass.read_command('db.databases', driver='sqlite').replace('\n', '')
    con = sqlite3.connect(sql_path)
    sql_stat = 'SELECT * FROM %s ' % cover_shapefile
    land_soil_shp_df = pd.read_sql_query(sql_stat, con, index_col='cat')
    land_cover_shp_codes = land_soil_shp_df[clc_field_name]
    land_cover_shp_codes = [np.int(i) for i in land_cover_shp_codes]
    soil_types_shp_codes = land_soil_shp_df[soil_field_name]
    soil_types_shp_codes = [i.lower() for i in soil_types_shp_codes]

    # create curve number column
    if 'CN' not in land_soil_shp_df.columns.values:
        grass.run_command('v.db.addcolumn', map=cover_shapefile, columns='CN int')

    # read lookup table with land cover and soil type references
    land_cover_tb_df = pd.read_excel(lookup_file_path, 'Cobertura CLC', header=0, index_col=0)
    soil_type_tb_df = pd.read_excel(lookup_file_path, 'Clasificacion Suelo', header=0, index_col=1)

    land_cover_references = [np.int(i) for i in land_cover_tb_df.index.values]
    soil_type_references = [i.lower() for i in soil_type_tb_df.index.values]

    soil_type_tb_df.index = soil_type_references

    # check if all shapefile covers and soil types are related to a CN in the lookup file
    cover_matches = False
    soil_matches = False
    for i, land_cover_code in enumerate(land_cover_shp_codes):
        if land_cover_code not in land_cover_references:
            cover_matches = True
            print("!%s cover wasn't found in cover type reference file!" % land_cover_code)
    for j, soil_type_code in enumerate(soil_types_shp_codes):
        if soil_type_code not in soil_type_references:
            soil_matches = True
            print("!%s soil wasn't found in soil type reference file!" % soil_type_code)

    # associate CN to specific land cover and soil type
    try:
        if not cover_matches and not soil_matches:
            for i in land_soil_shp_df.index.values:
                soil_code = land_soil_shp_df[soil_field_name][i].lower()
                cover_code = int(land_soil_shp_df[clc_field_name][i])
                soil_hydro_group = soil_type_tb_df['Grupo_Hidrologico'][soil_code]
                curve_number = land_cover_tb_df[soil_hydro_group][cover_code]
                grass.run_command('v.db.update', map=cover_shapefile, column='CN', where='cat=%s' % i,
                                  value=curve_number)

            grass.run_command('v.to.rast', input=cover_shapefile, output=curve_number_r, use='attr',
                              attribute_column='CN', overwrite=True)
            print('!CN estimation DONE!')
            print('Elapsed time = %0.2f seconds' % (time() - start_time))
    except TypeError:
        print('!CN estimation failed!')


def r_accumulation_weighting(flow_direction_r, flow_accumulation_r, input_r, output_weighted_r):
    """Raster weighting based on flow accumulation.

    Parameters
    ----------
    grass: object
        GRASS working environment (grass created with grass_environment)
    flow_direction_r: str
        Flow directions map [45 degree format e.g r.watershed direction]
    flow_accumulation_r: str
        Flow accumulation map [cells count]
    input_r: str
        Input map
    curve_number_weighted_r: str
        Weighted variable map by accumulation

    Returns
    -------
    Weighted map based on flow accumulation

    """
    # ==================================================================================================================
    # -- Estimate weighted variable
    # ==================================================================================================================
    grass.run_command('r.mask', raster=input_r)

    # get pixel resolution in square kilometers
    pixel_res = grass.raster_info(input_r)['ewres']
    pixel_area_sq_km = pixel_res ** 2 / 1000000

    # curve number weighted by each pixel area
    expr = '%s = %s * %0.10f' % ('tmp_var', input_r, pixel_area_sq_km)
    grass.run_command('r.mapcalc', expression=expr, overwrite=True)

    # accumulative area
    expr = '%s = %s * %0.10f' % ('tmp_areas', flow_accumulation_r, pixel_area_sq_km)
    grass.run_command('r.mapcalc', expression=expr, overwrite=True)

    # accumulate weighted curve number
    grass.run_command('r.accumulate', direction=flow_direction_r, format='45degree', weight='tmp_var',
                      accumulation='tmp_var_acc', overwrite=True)

    # get weighted curve number
    expr = '%s = %s / %s' % (output_weighted_r, 'tmp_var_acc', 'tmp_areas')
    grass.run_command('r.mapcalc', expression=expr, overwrite=True)

    # delete temporary raster and remove mask
    grass.run_command('g.remove', type='raster', name=['tmp_var_acc', 'tmp_var', 'tmp_areas'], flags='f')
    grass.run_command('r.mask', flags='r')


def r_regional_integration():
    """TODO: curve number by sub-basins."""


def v_extract_value(points_v, input_r, max_hydrology_file, out_sheetname='2. Runoff Coefficient', exclusion_file=None):
    """Extract raster value for each analysis point.

    TODO:Generalize r.what
    Parameters
    ----------
    points_v: pandas.DataFrame
        Points vector representing basins outlets
    curve_number_weighted_r: str
        Weighted Curve Number map
    max_hydrology_file: str
        Excel file containing hydrology analysis of maximum streamflows
    exclusion_file: str
        Excel file containing list of excluded points from hydrology analysis

    Returns
    -------
    Extract Curve Number to points table and maximum hydrology file

    """
    # load attribute table as dataframe
    sql_path = grass.read_command('db.databases', driver='sqlite').replace('\n', '')
    con = sqlite3.connect(sql_path)
    sql_stat = 'SELECT * FROM %s ' % points_v
    analysis_points_df = pd.read_sql_query(sql_stat, con, index_col='cat')

    # excluded points from analysis
    # if exclusion_list:
    #     exclusion_list = pd.read_excel(exclusion_file, header=0)['cat']
    #     cols = ['stream_' + str(int(i)) for i in analysis_points_df[basin_id_field].values
    #             if i not in exclusion_list.values]
    # else:
    #     cols = ['stream_' + str(int(i)) for i in analysis_points_df[basin_id_field].values]

    # ==================================================================================================================
    # -- Create dataframe to store values
    # ==================================================================================================================
    cn_df = pd.DataFrame(data=None, index=['CN_AMCII', 'CN_AMCI', 'CN_AMCIII'],
                         columns=analysis_points_df['SUB_BASIN'].values)

    # ==================================================================================================================
    # -- Read each CN value from map and export to hydrology file
    # ==================================================================================================================
    # excluded points from analysis
    if exclusion_file:
        exclusion_list = pd.read_excel(exclusion_file, header=0)['cat']

        for i in analysis_points_df.index.values:
            if analysis_points_df['value'][i] not in exclusion_list.values:
                east_coord = analysis_points_df['east'][i]
                north_coord = analysis_points_df['north'][i]
                curve_number = grass.read_command('r.what', map=input_r,
                                                  coordinates=[east_coord, north_coord])
                curve_number = round(float(curve_number.split('|')[3]), 0)

                cn_amc_i = round(4.2*curve_number/(10-0.058*curve_number), 0)
                cn_amc_iii = round(23*curve_number/(10+0.13*curve_number), 0)

                cn_df['stream_' + str(int(analysis_points_df['value'][i]))]['CN_AMCII'] = curve_number
                cn_df['stream_' + str(int(analysis_points_df['value'][i]))]['CN_AMCI'] = cn_amc_i
                cn_df['stream_' + str(int(analysis_points_df['value'][i]))]['CN_AMCIII'] = cn_amc_iii
    else:
        for i in analysis_points_df.index.values:
            east_coord = analysis_points_df['EAST'][i]
            north_coord = analysis_points_df['NORTH'][i]
            curve_number = grass.read_command('r.what', map=input_r,
                                              coordinates=[east_coord, north_coord])
            curve_number = round(float(curve_number.split('|')[3]), 0)

            cn_amc_i = round(4.2*curve_number/(10-0.058*curve_number), 0)
            cn_amc_iii = round(23*curve_number/(10+0.13*curve_number), 0)

            cn_df[analysis_points_df['SUB_BASIN'][i]]['CN_AMCII'] = curve_number
            cn_df[analysis_points_df['SUB_BASIN'][i]]['CN_AMCI'] = cn_amc_i
            cn_df[analysis_points_df['SUB_BASIN'][i]]['CN_AMCIII'] = cn_amc_iii

    # save to a new excel sheet
    book = load_workbook(max_hydrology_file)
    writer = pd.ExcelWriter(max_hydrology_file, engine='openpyxl')
    writer.book = book

    cn_df.to_excel(writer, sheet_name=out_sheetname)
    writer.save()
    writer.close()


def r_voronoi_interpolation(analysis_points, voronoi_polygons_v, voronoi_polygons_v_field, voronoi_polygons_r):
    """
    Calculate Voronoi polygons for IDF curves stations.

    Parameters
    ----------
    analysis_points: str
        Points vector containing analysis points [stations with IDF curves]
    voronoi_polygons_v: str
        Voronoi polygons vector
    voronoi_polygons_v_field: str
        Voronoi polygons vector field for coding into map
    voronoi_polygons_r: str
        Voronoi polygons map

    Returns
    -------
    Voronoi polygons vector and map (inside GRASS mapset)

    """
    grass.run_command('v.voronoi', input=analysis_points, output=voronoi_polygons_v, overwrite=True)
    grass.run_command('v.to.rast', input=voronoi_polygons_v, use='attr', attribute_column=voronoi_polygons_v_field,
                      output=voronoi_polygons_r, overwrite=True)


def basin_polygon_weighting(analysis_points, categories_r, max_hydrology_file, exclusion_file=None, out_sheetname=None):
    """
    Estimates percentage of each category intercepting each basin in region.

    Parameters
    ----------
    garray: object
        GRASS garray module
    categories_r: str
        Input categories map
    max_hydrology_file: str
        Excel file containing hydrology analysis of maximum streamflows [must have morphometry and CN]
    exclusion_file: str
        Excel file containing list of excluded points from hydrology analysis

    Returns
    -------
    Weighting of categories for each sub-basin [in hydrology file]

    """
    # load attribute table as dataframe
    sql_path = grass.read_command('db.databases', driver='sqlite').replace('\n', '')
    con = sqlite3.connect(sql_path)
    sql_stat = 'SELECT * FROM %s ' % analysis_points
    points_df = pd.read_sql_query(sql_stat, con, index_col='cat')

    # excluded points from analysis
    if exclusion_file:
        exclusion_list = pd.read_excel(exclusion_file, header=0)['cat']
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values if i not in exclusion_list.values]
    else:
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values]

    # ==================================================================================================================
    # -- Weight categories raster for each sub-basin
    # ==================================================================================================================
    # load categories map
    categories_raster = garray.array()
    categories_raster.read(categories_r)

    # categories dataframe
    categories_list = np.unique(categories_raster)
    categories_df = pd.DataFrame(data=np.zeros([len(categories_list), len(cols)]), index=categories_list, columns=cols)

    # weight categories in each sub-basin
    counter = 1
    for i in cols:
        basin_raster = []
        basin_raster = garray.array()
        basin_raster.read(str(i) + "_basin")

        is_basin = np.where(basin_raster != 0)
        basin_category_raster = categories_raster[is_basin]
        category_counts = np.unique(basin_category_raster, return_counts=True)
        category_total_cells = category_counts[1].sum()
        category_percents = category_counts[1] / category_total_cells
        for j in enumerate(category_percents):
            categories_df.loc[category_counts[0][j], i] = category_percents[j]

        print('%s of %s DONE' % (counter, len(cols)))
        counter += 1

    # ==================================================================================================================
    # -- Export time of concentration to hydrology file
    # ==================================================================================================================
    book = load_workbook(max_hydrology_file)
    writer = pd.ExcelWriter(max_hydrology_file, engine='openpyxl')
    writer.book = book

    categories_df.to_excel(writer, sheet_name=out_sheetname)
    writer.save()
    writer.close()
