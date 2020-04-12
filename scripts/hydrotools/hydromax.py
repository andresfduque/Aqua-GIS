#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Created by AndresD on 8/06/19.

Module tha integrates best of geomorphometric tools available in different GIS open source software

Features:
    + Time of concentration
    + IDF curves fit
    + Area reduction factor
    + Design storm
    + Unit hydrographs
    + Flow hydrographs

Pre-requisites:
    + GRASS python scripting library
    + WhiteBox Geospatial Data Analysis (pip installed)

@author:    Andres Felipe Duque Perez
Email:      aduquep@ingenevo.com.co

Credits:

Lindsay, J. B. (2016). Whitebox GAT: A case study in geomorphometric analysis. Computers & Geosciences, 95, 75-84.
http://dx.doi.org/10.1016/j.cageo.2016.07.003

Neteler, M., Bowman, M.H., Landa, M., Metz, M., 2012. GRASS GIS: A multi-purpose open source GIS. Environ Model Soft 31,
124–130.
https://doi.org/10.1016/j.envsoft.2011.11.014
"""


def time_concentration(analysis_points, max_hydrology_file, iqr_interval=1.0, amc='AMC II', exclusion_file=None):
    """
    Calculate time of concentration of the basins.

    https://www.researchgate.net/publication/316417342_Estimating_time_of_concentration_in_large_watersheds
    https://www.degruyter.com/downloadpdf/j/sjce.2016.24.issue-4/sjce-2016-0017/sjce-2016-0017.pdf
    https://www.researchgate.net/publication/305304627_Desempenho_de_Formulas_de_Tempo_de_Concentracao_em_Bacias_Urbanas_e_Rurais

    Parameters
    ----------
    grass: object
        GRASS working environment (grass created with grass_environment)
    analysis_points: str
        Points vector containing analysis points
    max_hydrology_file: str
        Excel file containing hydrology analysis of maximum streamflows [must have morphometry and CN]
    iqr_interval: float
        Acceptable range of deviation from median
    amc: str
        Hydrological previous conditions [default= II]
    exclusion_file: str
        Excel file containing list of excluded points from hydrology analysis

    Returns
    -------
    Curve number map (inside GRASS mapset)

    """
    # ==================================================================================================================
    # -- Main imports and preparations
    # ==================================================================================================================
    import sqlite3
    import numpy as np
    import pandas as pd
    from scipy.stats import iqr
    import grass.script as grass
    from openpyxl import load_workbook

    # load attribute table as dataframe
    sql_path = grass.read_command('db.databases', driver='sqlite').replace('\n', '')
    con = sqlite3.connect(sql_path)
    sql_stat = 'SELECT * FROM %s ' % analysis_points
    points_df = pd.read_sql_query(sql_stat, con, index_col='cat')

    # units conversions
    km_2_mi = 0.621371
    km_2_ft = 3280.84
    mm_ft_miles = 5280

    # excluded points from analysis
    if exclusion_file:
        exclusion_list = pd.read_excel(exclusion_file, header=0)['cat']
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values if i not in exclusion_list.values]
    else:
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values]

    # dataframe to store times of concentration
    concentration_time_df = pd.DataFrame(data=None, index=['Bransby-Williams', 'California Culverts Practice',
                                                           'Clark', 'Giandotti', 'Johnstone-Cross', 'Kirpich',
                                                           'Passini', 'Perez', 'Pilgrim-McDermott', 'Snyder',
                                                           'SCS Lag', 'Temez', 'Valencia-Zuluaga', 'Ven Te Chow',
                                                           'Ventura', 'Dooge', 'Corps Engineers', 'Carter', 'Picking',
                                                           'Mean', 'Median', 'Standard Deviation', 'IQR',
                                                           'Interval Amplitude', 'Upper Limit', 'Lower Limit',
                                                           'Selected'], columns=cols)

    # ==================================================================================================================
    # -- Get morphometric parameters of each basin and calculate times of concentration
    # ==================================================================================================================
    morphometric_df = pd.read_excel(max_hydrology_file, sheet_name='1. Morphometric Parameters', header=0, index_col=0)
    curve_number_df = pd.read_excel(max_hydrology_file, sheet_name='2. Runoff Coefficient', header=0, index_col=0)

    for i in morphometric_df.columns.values:
        # get morphometric parameters
        basin_area = morphometric_df[i]['7. Basin Area [km^2]']
        basin_distance_cg = morphometric_df[i]['11. Basin Distance to Center of Gravity [km]']
        basin_elevation_difference = morphometric_df[i]['14. Basin Elevation Difference [m]']
        basin_mean_slope = morphometric_df[i]['16. Basin Mean Slope [m/m]']
        mainchannel_length = morphometric_df[i]['19. Mainchannel Length [km]']
        mainchannel_mean_slope = morphometric_df[i]['20. Mainchannel Mean Slope [m/m]']

        # prevent zero channel slope
        if mainchannel_mean_slope <= 0.001:
            mainchannel_mean_slope = 0.001

        # prevent negative distance to centroid [when stream start matches with outlet point, very rare]
        if basin_distance_cg < 0:
            basin_distance_cg = 0.1

        # get Antecedent Moisture Condition [AMC]
        if amc == 'AMC II':
            curve_number = curve_number_df[i]['CN_AMCII']
        elif amc == 'AMC III':
            curve_number = curve_number_df[i]['CN_AMCIII']
        elif amc == 'AMC I':
            curve_number = curve_number_df[i]['CN_AMCI']

        # Bransby-Williams [Department of Transport and Main Roads, 2010] - CONFIRMED in 3 sources
        tc_bransby = 14.46 * mainchannel_length * (basin_area ** -0.1) * mainchannel_mean_slope ** -0.2

        # California Culvert Practice (1955) - Li and Chibber (2008) - CONFIRMED in 3 sources
        tc_california_culverts_practice = (0.87075 * (mainchannel_length ** 3) /
                                           basin_elevation_difference) ** 0.385
        tc_california_culverts_practice = tc_california_culverts_practice * 60  # convert to minutes

        # Clark (1945) - CONFIRMED in 1 source
        tc_clark = 0.335 * (basin_area / basin_mean_slope ** 0.5) ** 0.593
        tc_clark = tc_clark * 60                        # convert to minutes

        # Giandotti (1934) - Del Giudice et al. (2012) - Shariﬁ and Hosseini (2011) - CONFIRMED in 3 sources
        tc_giandotti = (4 * basin_area ** 0.5 + 1.5 * mainchannel_length) / \
            (0.8 * basin_elevation_difference ** 0.5)
        tc_giandotti = tc_giandotti * 60                # convert to minutes

        # Johnstone and Cross (1949) - CONFIRMED in 1 source
        # SILVEIRA, A.L.L. Performance of time of concentration formulas in urban and rural basins.
        # Revista Brasileira de Recursos Hídricos, v.10, p.5-23, 2005
        tc_johnstone_cross = 5 * ((mainchannel_length * km_2_mi) / (basin_mean_slope * mm_ft_miles) ** 0.5) ** 0.5
        tc_johnstone_cross = tc_johnstone_cross * 60    # convert to minutes

        # Kirpich (1940) - Li and Chibber (2008) - CONFIRMED in 1 source
        # SILVEIRA, A.L.L. Performance of time of concentration formulas in urban and rural basins.
        # Revista Brasileira de Recursos Hídricos, v.10, p.5-23, 2005
        tc_kirpich = 0.0078 * ((mainchannel_length * km_2_ft) ** 0.77) * mainchannel_mean_slope ** -0.385

        # Passini [Department of Transport and Main Roads, 2010] - CONFIRMED in 1 source
        # SILVEIRA, A.L.L. Performance of time of concentration formulas in urban and rural basins.
        # Revista Brasileira de Recursos Hídricos, v.10, p.5-23, 2005
        tc_passini = 0.108 * ((basin_area * mainchannel_length) ** (1/3)) * mainchannel_mean_slope ** -0.5
        tc_passini = tc_passini * 60                    # convert to minutes

        # Perez, O. Determinación del tiempo de concentración para estimar la avenida de diseño. Ingeniería civil
        # ( La Habana ), Cuba, 1985 V36 N1 ene-feb p 40-53. 1985.
        tc_perez = mainchannel_length / (72 * (basin_elevation_difference / 1000 / mainchannel_length) ** 0.6)
        tc_perez = tc_perez * 60                        # convert to minutes

        # Pilgrim D.H & McDermott, G.E. Desgin floods for small rural catchments in Eastern New South Wales .
        # Civ. Eng. Trans. Inst. Engs. Aust. Vol CE24.pp. 226-234. 1982. - CONFIRMED in 3 sources
        tc_pilgrim = 0.76 * basin_area ** 0.38
        tc_pilgrim = tc_pilgrim * 60                    # convert to minutes

        # Snyder, F.F. Synthetic Unit Graphs. Transaction of the American Geophysical Union . Estados Unidos. 1938.
        # http://www.engr.colostate.edu/~ramirez/ce_old/classes/ce522_ramirez/snyder/snyder_uh.htm
        tc_snyder = 1.2 * (mainchannel_length * basin_distance_cg) ** 0.3
        tc_snyder = tc_snyder * 60                      # convert to minutes

        # NRCS (1997) - CONFIRMED in all sources
        tc_scs_lag = (100 * ((mainchannel_length*km_2_ft) ** 0.8) * (1000 / curve_number - 9) ** 0.7) / \
            (1900*(mainchannel_mean_slope * 100) ** 0.5)

        # Temez (1978) - CONFIRMED in 2 sources
        # TEMEZ, J.R. Extended and improved Rational Method. Proc. XXIV Congress, Madrid, España.
        # Vol. A. pp 33-40. 1991
        tc_temez = 0.3 * (mainchannel_length / mainchannel_mean_slope ** 0.25) ** 0.76
        tc_temez = tc_temez * 60                        # convert to minutes

        # Valencia, C & Zuluaga, O. Estudio preliminar del tiempo de concentración en algunas cuencas de Antioquia.
        # Trabajo de grado. Universidad Nacional de Colombia, sede Medellín. 1981.
        tc_valencia_zuluaga = 1.7694 * (basin_area ** 0.325) * (mainchannel_length ** -0.096) * \
            ((mainchannel_mean_slope * 100) ** -0.29)
        tc_valencia_zuluaga = tc_valencia_zuluaga * 60

        # Ven Te Chow (1962)- CONFIRMED in all sources
        tc_chow = 0.1602 * mainchannel_length ** 0.64 * mainchannel_mean_slope ** -0.32
        tc_chow = tc_chow * 60

        # Ventura [Guermond, 2008] - CONFIRMED in all source
        tc_ventura = 7.62 * (basin_area / mainchannel_mean_slope) ** 0.5

        # Simas-Hawkins - CONFIRMED in 1 source - [DEPRECATED]
        # SILVEIRA, A.L.L. Performance of time of concentration formulas in urban and rural basins.
        # Revista Brasileira de Recursos Hídricos, v.10, p.5-23, 2005
        tc_simas = 0.322 * basin_area ** 0.594 * mainchannel_length ** -0.594 * mainchannel_mean_slope ** - 0.15 * \
            (25400 / curve_number - 254) ** 0.313
        tc_simas = tc_simas * 60

        # Dooge - CONFIRMED in 1 source
        # SILVEIRA, A.L.L. Performance of time of concentration formulas in urban and rural basins.
        # Revista Brasileira de Recursos Hídricos, v.10, p.5-23, 2005
        tc_dooge = 0.365 * basin_area ** 0.41 * mainchannel_mean_slope ** -0.17
        tc_dooge = tc_dooge * 60

        # Corps of engineers - CONFIRMED in 1 source
        # SILVEIRA, A.L.L. Performance of time of concentration formulas in urban and rural basins.
        # Revista Brasileira de Recursos Hídricos, v.10, p.5-23, 2005
        tc_corps = 0.191 * mainchannel_length ** 0.76 * mainchannel_mean_slope ** -0.19
        tc_corps = tc_corps * 60

        # Carter - CONFIRMED in 1 source
        # SILVEIRA, A.L.L. Performance of time of concentration formulas in urban and rural basins.
        # Revista Brasileira de Recursos Hídricos, v.10, p.5-23, 2005
        tc_carter = 0.0977 * mainchannel_length ** 0.6 * mainchannel_mean_slope ** -0.3
        tc_carter = tc_carter * 60

        # Picking - CONFIRMED in 1 source
        # SILVEIRA, A.L.L. Performance of time of concentration formulas in urban and rural basins.
        # Revista Brasileira de Recursos Hídricos, v.10, p.5-23, 2005
        tc_picking = 0.0883 * mainchannel_length ** 0.6667 * mainchannel_mean_slope ** -0.3333
        tc_picking = tc_picking * 60

        # times of concentration vector
        tc_vector = np.array([tc_bransby, tc_california_culverts_practice, tc_clark, tc_giandotti,
                              tc_johnstone_cross, tc_kirpich, tc_passini, tc_perez, tc_pilgrim, tc_snyder,
                              tc_scs_lag, tc_temez, tc_valencia_zuluaga, tc_chow, tc_ventura, tc_dooge, tc_corps,
                              tc_carter, tc_picking])

        # times of concentration metrics uncorrected vector
        tc_median = np.median(tc_vector)
        tc_std = np.std(tc_vector)

        # correct vector eliminating outliers
        tc_vector_corrected = tc_vector[np.where((tc_vector >= tc_median - 2 * tc_std) &
                                                 (tc_vector <= tc_median + 2 * tc_std))]
        tc_mean = np.mean(tc_vector_corrected)
        tc_median = np.median(tc_vector_corrected)
        tc_std = np.std(tc_vector_corrected)

        tc_iqr = iqr(tc_vector_corrected)
        upper_limit = tc_median + iqr_interval * tc_iqr / 2
        lower_limit = tc_median - iqr_interval * tc_iqr / 2
        if lower_limit < 0:
            lower_limit = 0

        # select time of concentration as the average of data between the range with an amplitude equal to the IQR
        # and centered in the median
        tc_selected = 0     # minutes
        c = 0
        for j in tc_vector_corrected:
            if lower_limit <= j <= upper_limit:
                tc_selected += j
                c += 1

        # if concentration time is less than 5 minutes make it 5 minutes
        tc_selected = tc_selected / c
        if tc_selected < 5:
            tc_selected = 5

        tc_selected = round(tc_selected, 0)

        # ==============================================================================================================
        # -- Fill time of concentration dataframe and points vector
        # ==============================================================================================================
        if exclusion_file:
            if not int(i[7:]) in exclusion_list.values:
                concentration_time_df[i] = np.append(tc_vector, [tc_mean, tc_median, tc_std, tc_iqr, iqr_interval,
                                                                 upper_limit, lower_limit, tc_selected])
        else:
            concentration_time_df[i] = np.append(tc_vector, [tc_mean, tc_median, tc_std, tc_iqr, iqr_interval,
                                                             upper_limit, lower_limit, tc_selected])

    # ==================================================================================================================
    # -- Export time of concentration to hydrology file
    # ==================================================================================================================
    book = load_workbook(max_hydrology_file)
    writer = pd.ExcelWriter(max_hydrology_file, engine='openpyxl')
    writer.book = book

    concentration_time_df.to_excel(writer, sheet_name='3. Concentration Time')
    writer.save()
    writer.close()


def idf_curves_fit(max_hydrology_file, idf_file=None, sheet='5A. IDF Parameters Present'):
    """
    Calculate IDF curves parameters.

        Equation Type[1] = (a * TR ^ m) / (b + D) ^ n
        Equation Type[2] = (a * TR ^ m) / (b + D ^ n)

    Parameters
    ----------
    idf_file: str
        Excel file containing IDF values
    max_hydrology_file: str
        Excel file containing hydrology analysis of maximum streamflows [must have morphometry and CN]

    Returns
    -------
    IDF curves parameters [in hydrology file]

    """
    # ==================================================================================================================
    # -- Main imports and preparations
    # ==================================================================================================================
    import pandas as pd
    from openpyxl import load_workbook
    from scipy.optimize import minimize

    # ==================================================================================================================
    # -- Fit IDF Curve
    # ==================================================================================================================
    # IDF function type I, returning the sum of squared errors (SSE)
    def func1(par, res):
        p1 = par[0] * res.index.values ** par[1]
        p2 = (res.columns.values.astype(float) + par[2]) ** par[3]
        f = pd.DataFrame([p / p2 for p in p1], index=res.index, columns=res.columns)
        error_tot_q = ((f - res) ** 2).sum(axis=1).sum()
        return error_tot_q

    # IDF function type II, returning the sum of squared errors (SSE)
    def func2(par, res):
        p1 = par[0] * res.index.values ** par[1]
        p2 = (res.columns.values.astype(float) ** par[3] + par[2])
        f = pd.DataFrame([p / p2 for p in p1], index=res.index, columns=res.columns)
        error_tot_q = ((f - res) ** 2).sum(axis=1).sum()
        return error_tot_q

    # create dataframe to store regression values
    idf_parameters_df = pd.DataFrame(data=None, index=['Eq. type', 'Return period', 'a', 'm', 'b', 'n'])

    if idf_file:
        xl = pd.ExcelFile(idf_file)
        equation_type = pd.read_excel(idf_file, sheet_name='Equation Type', header=0, index_col='Gauge Station')

        for i in xl.sheet_names:
            if i != 'Equation Type':
                # copy your rainfall intensities table from excel with column headers with return period in years and
                # row names with rainfall durations in minutes
                df_int = pd.read_excel(idf_file, sheet_name=i, header=0, index_col='Duracion')
                df_int = df_int.transpose()

                # initial guess
                param1 = [5000, 0.1, 10, 0.9]

                # find parameters
                eq_type = equation_type['Equation Type'][i]
                if eq_type == 1:
                    res2 = minimize(func1, param1, args=(df_int,), method='Nelder-Mead')
                elif eq_type == 2:
                    res2 = minimize(func2, param1, args=(df_int,), method='Nelder-Mead')

                # store in dataframe
                for j in df_int.index.values:
                    idf_parameters_df[i + '-Tr' + str(int(j))] = [eq_type, j, res2.x[0], res2.x[1], res2.x[2],
                                                                  res2.x[3]]

    # save to a new excel sheet
    book = load_workbook(max_hydrology_file)
    writer = pd.ExcelWriter(max_hydrology_file, engine='openpyxl')
    writer.book = book

    idf_parameters_df.to_excel(writer, sheet_name=sheet)
    writer.save()
    writer.close()


def area_reduction_factor(analysis_points, max_hydrology_file, exclusion_file=None):
    """
    Area reduction factor.

    Parameters
    ----------
    analysis_points: str
        Analysis points vector
    max_hydrology_file: str
        Excel file containing hydrology analysis of maximum streamflows [must have morphometry and CN]
    exclusion_file: str
        Excel file containing list of excluded points from hydrology analysis

    Returns
    -------
    Area Reduction Factor [in hydrology file]

    """
    # ==================================================================================================================
    # -- Main imports and preparations
    # ==================================================================================================================
    import sqlite3
    import numpy as np
    import pandas as pd
    import grass.script as grass
    from openpyxl import load_workbook
    import grass.script.array as garray

    # load attribute table as dataframe
    sql_path = grass.read_command('db.databases', driver='sqlite').replace('\n', '')
    con = sqlite3.connect(sql_path)
    sql_stat = 'SELECT * FROM %s ' % analysis_points
    points_df = pd.read_sql_query(sql_stat, con, index_col='cat')

    # excluded points from analysis
    if exclusion_file:
        exclusion_list = pd.read_excel(exclusion_file, header=0)['cat']
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values if i not in exclusion_list.values]
    else:
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values]

    arf_df = pd.DataFrame(data=None, index=['arf_temez'], columns=cols)

    # ==================================================================================================================
    # -- Estimate Areal Reduction Factor
    # ==================================================================================================================
    concentration_time_df = pd.read_excel(max_hydrology_file, sheet_name='1. Morphometric Parameters', header=0,
                                          index_col=0)

    for i in cols:
        area = concentration_time_df[i]['7. Basin Area [km^2]']
        if area <= 1:
            arf = 1
        else:
            arf = 1 - np.log10(area) / 15

        arf_df.loc['arf_temez', i] = arf

    # ==================================================================================================================
    # -- Export area reduction factor to hydrology file
    # ==================================================================================================================
    book = load_workbook(max_hydrology_file)
    writer = pd.ExcelWriter(max_hydrology_file, engine='openpyxl')
    writer.book = book

    arf_df.to_excel(writer, sheet_name='6. ARF')
    writer.save()
    writer.close()


def design_storm(analysis_points, max_hydrology_file, exclusion_file=None, scenario='only_present', intervals=20):
    """
    Get precipitation intensities.

    Parameters
    ----------
    analysis_points: str
        Analysis points vector
    max_hydrology_file: str
        Excel file containing hydrology analysis of maximum streamflows [must have morphometry and CN]
    exclusion_file: str
        Excel file containing list of excluded points from hydrology analysis
    scenario: str
        Precipitation scenarios [only present, only_future, present-future]

    Returns
    -------
    Area Reduction Factor [in hydrology file]

    """
    # ==================================================================================================================
    # -- Main imports and preparations
    # ==================================================================================================================
    import sqlite3
    import numpy as np
    import pandas as pd
    import grass.script as grass
    from openpyxl import load_workbook
    import grass.script.array as garray

    # load attribute table as dataframe
    sql_path = grass.read_command('db.databases', driver='sqlite').replace('\n', '')
    con = sqlite3.connect(sql_path)
    sql_stat = 'SELECT * FROM %s ' % analysis_points
    points_df = pd.read_sql_query(sql_stat, con, index_col='cat')

    # excluded points from analysis
    if exclusion_file:
        exclusion_list = pd.read_excel(exclusion_file, header=0)['cat']
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values if i not in exclusion_list.values]
    else:
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values]

    # ==================================================================================================================
    # -- Get intensities for each IDF curve
    # ==================================================================================================================
    # basin info dataframes
    arf_df = pd.read_excel(max_hydrology_file, sheet_name='6. ARF', header=0, index_col=0)
    voronoi_df = pd.read_excel(max_hydrology_file, sheet_name='4. Voronoi Percentages', header=0, index_col=0)
    curve_number_df = pd.read_excel(max_hydrology_file, sheet_name='2. Runoff Coefficient', header=0, index_col=0)
    morphometric_df = pd.read_excel(max_hydrology_file, sheet_name='1. Morphometric Parameters', header=0, index_col=0)
    concentration_time_df = pd.read_excel(max_hydrology_file, sheet_name='3. Concentration Time', header=0, index_col=0)

    # depending on scenario
    if scenario == 'only_present':
        idf_parameters_df = pd.read_excel(max_hydrology_file, sheet_name='5A. IDF Parameters Present', header=0,
                                          index_col=0)
        idf_stations_indexes = idf_parameters_df.columns.values
        idf_stations_unique = [i[:7] for i in idf_parameters_df.columns.values]
        idf_stations_unique = np.unique(idf_stations_unique)
        scenario_code = 'P'
    elif scenario == 'only_future':
        idf_parameters_df = pd.read_excel(max_hydrology_file, sheet_name='5B. IDF Parameters Future', header=0,
                                          index_col=0)
        idf_stations_indexes = idf_parameters_df.columns.values
        idf_stations_unique = [i[:7] for i in idf_parameters_df.columns.values]
        idf_stations_unique = np.unique(idf_stations_unique)
        scenario_code = 'F'

    # dataframe to store design storm intensity [mm/h] and depth [mm]
    depths_df = pd.DataFrame(data=None, index=idf_stations_indexes, columns=cols)
    intensities_df = pd.DataFrame(data=None, index=idf_stations_indexes, columns=cols)

    # dataframes to store design storm time distribution
    intervals_index = ['Interval %s' % i for i in range(1, intervals + 1)]
    storm_intervals = []
    return_periods = ['Tr2', 'Tr5', 'Tr10', 'Tr25', 'Tr50', 'Tr100', 'Tr200', 'Tr500']
    for i in return_periods:
        for j in intervals_index:
            storm_intervals.append(j + ' - ' + i)

    depth_distribution_df = pd.DataFrame(data=np.zeros([len(storm_intervals), len(cols)]), index=storm_intervals,
                                         columns=cols)
    depth_cumm_distribution_df = pd.DataFrame(data=np.zeros([len(storm_intervals), len(cols)]), index=storm_intervals,
                                              columns=cols)
    effective_depth_distribution_df = pd.DataFrame(data=np.zeros([len(storm_intervals), len(cols)]),
                                                   index=storm_intervals, columns=cols)
    effective_cumm_depth_distribution_df = pd.DataFrame(data=np.zeros([len(storm_intervals), len(cols)]),
                                                        index=storm_intervals, columns=cols)
    time_intervals_df = pd.DataFrame(data=np.zeros([intervals, len(cols)]), index=intervals_index, columns=cols)

    # calculate intensity, depth and depth distribution
    counter = 1
    for i in cols:
        area = morphometric_df[i]['7. Basin Area [km^2]']
        tc = concentration_time_df[i]['Selected']
        for j in idf_stations_indexes:
            a = idf_parameters_df[j]['a']
            m = idf_parameters_df[j]['m']
            b = idf_parameters_df[j]['b']
            n = idf_parameters_df[j]['n']
            tr = float(j.split('Tr')[-1])
            equation_type = idf_parameters_df[j]['Eq. type']
            if equation_type == 1:
                intensity = (a * tr ** m) / (b + tc) ** n
            elif equation_type == 2:
                intensity = (a * tr ** m) / (b + tc ** n)

            # weight intensity by IDF Voronoi percentage
            idf_station_id = int(j[:8])
            voronoi_factor = voronoi_df[i][idf_station_id]
            intensities_df.loc[j, i] = intensity * voronoi_factor

            # correct by ARF
            arf = arf_df[i]['arf_temez']
            depths_df.loc[j, i] = intensity * tc / 60 * voronoi_factor * arf

            # get depth time distribution
            tc_arange = np.arange(tc / intervals, tc + tc / intervals, tc / intervals)
            if len(tc_arange) > intervals:
                tc_arange = tc_arange[0:intervals]
            time_intervals_df.loc[intervals_index, i] = tc_arange
            depth_arange = np.zeros([len(tc_arange)])
            cumm_depth_arange = np.zeros([len(tc_arange)])

            for k in range(0, len(tc_arange)):
                if equation_type == 1:
                    d = tc_arange[k]
                    intensity = (a * tr ** m) / (b + d) ** n
                    depth = intensity * d / 60 * arf
                elif equation_type == 2:
                    d = tc_arange[k]
                    intensity = (a * tr ** m) / (b + d ** n)
                    depth = intensity * d / 60 * arf

                cumm_depth_arange[k] = depth
                if k < 1:
                    depth_arange[k] = depth
                else:
                    depth_arange[k] = depth - cumm_depth_arange[k - 1]

            # sort depth vector
            depth_arange_sorted = np.sort(depth_arange)
            depth_arange_sorted = depth_arange_sorted[::-1]

            # arange for alternating blocks method
            depth_blocks = np.zeros([len(tc_arange)])
            y = int(len(tc_arange) / 2)
            depth_blocks[y - 1] = depth_arange_sorted[0]
            for k in range(1, int(len(tc_arange) / 2) + 1):
                if k <= int(len(tc_arange) / 2 - 1):
                    depth_blocks[y - 1 + k] = depth_arange_sorted[(2 * k) - 1]
                    depth_blocks[y - 1 - k] = depth_arange_sorted[2 * k]
                else:
                    depth_blocks[y - 1 + k] = depth_arange_sorted[(2 * k) - 1]

            # update dataframe
            cumm_depth = 0
            cn = curve_number_df[i]['CN_AMCIII']
            for k in range(1, intervals + 1):
                interval_name = 'Interval %s' % k
                depth = depth_blocks[k - 1] * voronoi_factor
                aux_depth = depth_distribution_df[i][interval_name + ' - Tr' + str(int(tr))]
                cumm_depth += depth
                aux_cumm_depth = depth_cumm_distribution_df[i][interval_name + ' - Tr' + str(int(tr))]
                depth_distribution_df.loc[interval_name + ' - Tr' + str(int(tr)), i] = depth + aux_depth
                depth_cumm_distribution_df.loc[interval_name + ' - Tr' + str(int(tr)), i] = cumm_depth + aux_cumm_depth

                storage = (25400 - 254 * cn) / cn
                initial_abstraction = 0.05 * storage

                if cumm_depth <= initial_abstraction:
                    effective_cumm_depth = 0
                else:
                    effective_cumm_depth = (cumm_depth - initial_abstraction) ** 2 / \
                        (cumm_depth - initial_abstraction + storage)

                aux_effective_cumm_depth = effective_cumm_depth_distribution_df[i][interval_name + ' - Tr' +
                                                                                   str(int(tr))]
                effective_cumm_depth_distribution_df.loc[interval_name + ' - Tr' + str(int(tr)), i] = \
                    effective_cumm_depth + aux_effective_cumm_depth

            for k in range(1, intervals + 1):
                interval_name = 'Interval %s' % k
                if k > 1:
                    interval_name_aux = 'Interval %s' % (k - 1)
                    effective_cumm_depth = effective_cumm_depth_distribution_df[i][interval_name + ' - Tr' +
                                                                                   str(int(tr))]
                    aux_effective_cumm_depth = effective_cumm_depth_distribution_df[i][interval_name_aux + ' - Tr' +
                                                                                       str(int(tr))]
                    effective_depth = effective_cumm_depth - aux_effective_cumm_depth
                else:
                    effective_depth = 0

                effective_depth_distribution_df.loc[interval_name + ' - Tr' + str(int(tr)), i] = effective_depth

                # new figure
                # plt.bar(range(len(depth_blocks)), depth_blocks)
                # fig, ax = plt.subplots(figsize=[10, 6])
                # plt.show()

            # index = ['Interval %s' % i + ' - Tr' + str(int(tr)) for i in range(1, len(tc_arange) + 1)]
            # depth_distribution_df.loc['Interval - Tr' + str(tr), i] =
        print(str(counter) + ' of ' + str(len(cols)))
        counter += 1
    # ==================================================================================================================
    # -- Export area reduction factor to hydrology file
    # ==================================================================================================================
    book = load_workbook(max_hydrology_file)
    writer = pd.ExcelWriter(max_hydrology_file, engine='openpyxl')
    writer.book = book

    intensities_df.to_excel(writer, sheet_name='7. P Intensities')
    depths_df.to_excel(writer, sheet_name='8A. P Depth')
    time_intervals_df.to_excel(writer, sheet_name='8B. P Depth Intervals')
    depth_distribution_df.to_excel(writer, sheet_name='8C. P Depth Dist')
    depth_cumm_distribution_df.to_excel(writer, sheet_name='8D. P Depth Accum Dist')
    effective_cumm_depth_distribution_df.to_excel(writer, sheet_name='8E. P Efect Depth Accum Dist')
    effective_depth_distribution_df.to_excel(writer, sheet_name='8F. P Efect Depth Dist')

    writer.save()
    writer.close()


def unit_hydrographs(analysis_points, max_hydrology_file, exclusion_file=None, intervals=100):
    """
    Estimate unit Hydrographs.

    Parameters
    ----------
    analysis_points: str
        Analysis points vector
    max_hydrology_file: str
        Excel file containing hydrology analysis of maximum streamflows [must have morphometry and CN]
    exclusion_file: str
        Excel file containing list of excluded points from hydrology analysis
    scenario: str
        Precipitation scenarios [only present, only_future, present-future]
    intervals: int
        Number of periods to build the unit hydrograph

    Returns
    -------
    Williams & Hann, Snyder and NRCS unit hydrograph

    """
    # ==================================================================================================================
    # -- Main imports and preparations
    # ==================================================================================================================
    import sqlite3
    import numpy as np
    import pandas as pd
    import grass.script as grass
    from openpyxl import load_workbook
    import grass.script.array as garray

    # load attribute table as dataframe
    sql_path = grass.read_command('db.databases', driver='sqlite').replace('\n', '')
    con = sqlite3.connect(sql_path)
    sql_stat = 'SELECT * FROM %s ' % analysis_points
    points_df = pd.read_sql_query(sql_stat, con, index_col='cat')

    # excluded points from analysis
    if exclusion_file:
        exclusion_list = pd.read_excel(exclusion_file, header=0)['cat']
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values if i not in exclusion_list.values]
    else:
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values]

    # ==================================================================================================================
    # -- Base info and new Dataframes
    # ==================================================================================================================
    # base info dataframes
    morphometric_df = pd.read_excel(max_hydrology_file, sheet_name='1. Morphometric Parameters', header=0, index_col=0)
    concentration_time_df = pd.read_excel(max_hydrology_file, sheet_name='3. Concentration Time', header=0, index_col=0)
    time_intervals_df = pd.read_excel(max_hydrology_file, sheet_name='8B. P Depth Intervals', header=0, index_col=0)
    concentration_time = pd.read_excel(max_hydrology_file, sheet_name='3. Concentration Time', header=0, index_col=0)

    # dataframe to store design storm intensity [mm/h] and depth [mm]
    intervals_index = ['Interval %s' % i for i in range(1, intervals + 1)]
    hydrograph_time_intervals_df = pd.DataFrame(data=np.zeros([intervals, len(cols)]), index=intervals_index,
                                                columns=cols)
    w_h_unit_hydrographs_df = pd.DataFrame(data=np.zeros([intervals, len(cols)]), index=intervals_index,
                                           columns=cols)
    snyder_unit_hydrographs_df = pd.DataFrame(data=np.zeros([intervals, len(cols)]), index=intervals_index,
                                              columns=cols)
    nrcs_unit_hydrographs_df = pd.DataFrame(data=np.zeros([intervals, len(cols)]), index=intervals_index,
                                            columns=cols)

    # ==================================================================================================================
    # -- Calculate Hydrographs [W & H, Snyder, NRCS]
    # ==================================================================================================================
    counter = 1
    for i in cols:
        # time intervals
        storm_time_intervals = time_intervals_df[i].values
        interval_length = storm_time_intervals[1] - storm_time_intervals[0]
        hydrographs_intervals = np.arange(interval_length, 101 * interval_length, interval_length)
        if len(hydrographs_intervals) > intervals:
            hydrographs_intervals = hydrographs_intervals[:100]
        hydrograph_time_intervals_df.loc[intervals_index, i] = hydrographs_intervals

        # ==============================================================================================================
        # -- Williams & Hann unit hydrograph
        # ==============================================================================================================
        # required morphometric parameters
        basin_area = morphometric_df[i]['7. Basin Area [km^2]']
        basin_width = morphometric_df[i]['10. Basin Width [km]']
        basin_length = morphometric_df[i]['9. Basin Max Length [km]']
        mainchannel_slope = morphometric_df[i]['20. Mainchannel Mean Slope [m/m]']

        if mainchannel_slope <= 0.01:
            mainchannel_slope = 0.01

        mainchannel_slope_ft_mi = mainchannel_slope * 5280
        basin_area_miles = basin_area * 0.386102

        # form parameters
        k = 27 * (basin_area_miles ** 0.231) * (mainchannel_slope_ft_mi ** -0.777) * \
            (basin_length / basin_width) ** 0.124
        Tp = 4.63 * basin_area_miles ** 0.422 * mainchannel_slope_ft_mi ** -0.48 * (basin_length / basin_width) ** 0.133
        n = 1 + (1 / (2 * (k/Tp)) + (1 / (4 * (k/Tp) ** 2) + 1 / (k / Tp)) ** 0.5) ** 2

        to = Tp * (1 + 1 / (n - 1) ** 0.5)
        t1 = to + 2 * k

        if n <= 1.27:
            B = 7527.3397824 * n ** 3 - 28318.9594289 * n ** 2 + 35633.3593146 * n - 14987.3755403
        else:
            B = -0.0053450748 * n ** 5 + 0.1120132788 * n ** 4 - 0.1735995123 * n ** 3 - 12.79458518 * n ** 2 + \
                163.3452557299 * n - 85.1829993108

        Up = B * basin_area_miles / Tp * 0.0283168 * (1 / 25.4)     # ft3/s to m3/s and inches to mm

        # vector to store unit hydrograph
        w_h_unit_hydrograph = np.zeros([len(hydrographs_intervals)])

        # calculate hydrograph
        section = 0
        for j in range(0, len(hydrographs_intervals)):
            time = hydrographs_intervals[j] / 60
            if time < to:
                w_h_unit_hydrograph[j] = Up * (time / Tp) ** (n - 1) * np.exp((1 - n) * (time / Tp - 1))
            elif time >= to and time <= t1:
                if section == 0:
                    Uo = w_h_unit_hydrograph[j - 1]
                    section = 1
                w_h_unit_hydrograph[j] = Uo * np.exp((to - time) / k)
            elif time > t1:
                if section == 1:
                    U1 = w_h_unit_hydrograph[j - 1]
                    section = 2
                w_h_unit_hydrograph[j] = U1 * np.exp((t1 - time) / (3 * k))

        # update hydrographs dataframe
        w_h_unit_hydrographs_df.loc[intervals_index, i] = w_h_unit_hydrograph

        import matplotlib.pyplot as plt

        # plt.scatter(hydrographs_intervals, w_h_unit_hydrograph)
        # plt.show()

        # ==============================================================================================================
        # -- Snyder unit hydrograph
        # ==============================================================================================================
        # required morphometric parameters
        basin_cg = morphometric_df[i]['11. Basin Distance to Center of Gravity [km]']
        mainchannel_length = morphometric_df[i]['19. Mainchannel Length [km]']
        concentration_time = concentration_time_df[i]['Selected']

        if mainchannel_length <= 0.1:
            mainchannel_length = 0.1

        if basin_cg <= 0.01:
            basin_cg = 0.01

        # form parameters
        Ct = 0.420
        Cp = 0.80
        tp = 0.75 * Ct * (basin_cg * mainchannel_length) ** 0.3
        if tp > 0.6 * concentration_time / 60:
            tp = 0.6 * concentration_time / 60
        tr = tp / 5.5
        tR = interval_length / 60
        tpR = tp - (tr - tR) / 4

        # peak flows
        qp = 2.75 * Cp * basin_area / tp / 10
        qpR = qp * tp / tpR

        # times
        tb = 5.56 * basin_area / (qpR * 10)
        w50 = 2.14 / ((qp / basin_area) * 10) ** 1.08
        w75 = 1.22 / ((qp / basin_area) * 10) ** 1.08

        # unit hydrograph
        unit_hydrograph_snyder = np.zeros([7, 2])    # time, flow
        unit_hydrograph_snyder[0, 0] = 0
        unit_hydrograph_snyder[1, 0] = tp + tR / 2 - w50/3
        unit_hydrograph_snyder[2, 0] = tp + tR / 2 - w75/3
        unit_hydrograph_snyder[3, 0] = tp + tR / 2
        unit_hydrograph_snyder[4, 0] = tp + tR / 2 + w75 * 2 / 3
        unit_hydrograph_snyder[5, 0] = tp + tR / 2 + w50 * 2 / 3
        unit_hydrograph_snyder[6, 0] = tb

        unit_hydrograph_snyder[0, 1] = 0
        unit_hydrograph_snyder[1, 1] = 0.50 * qpR
        unit_hydrograph_snyder[2, 1] = 0.75 * qpR
        unit_hydrograph_snyder[3, 1] = qpR
        unit_hydrograph_snyder[4, 1] = 0.75 * qpR
        unit_hydrograph_snyder[5, 1] = 0.50 * qpR
        unit_hydrograph_snyder[6, 1] = 0

        # vector to store unit hydrograph
        snyder_unit_hydrograph = np.zeros([len(hydrographs_intervals)])

        # calculate hydrograph
        section = 0
        for j in range(0, len(hydrographs_intervals)):
            time = hydrographs_intervals[j] / 60
            snyder_unit_hydrograph[j] = np.interp(time, unit_hydrograph_snyder[:, 0], unit_hydrograph_snyder[:, 1])

        # update hydrographs dataframe
        snyder_unit_hydrographs_df.loc[intervals_index, i] = snyder_unit_hydrograph

        # plt.plot(hydrographs_intervals, snyder_unit_hydrograph)
        # plt.show()

        # ==============================================================================================================
        # -- NRCS unit hydrograph
        # ==============================================================================================================
        # form parameters
        tlag = 0.6 * concentration_time / 60
        tp = interval_length / 60 / 2 + tlag
        tb = 2.67 * tp

        # peak flows
        qp = 0.0208 * 100 * basin_area / tp / 10

        # unit hydrograph
        unit_hydrograph_nrcs = np.zeros([3, 2])    # time, flow
        unit_hydrograph_nrcs[0, 0] = 0
        unit_hydrograph_nrcs[1, 0] = tp
        unit_hydrograph_nrcs[2, 0] = tb

        unit_hydrograph_nrcs[0, 1] = 0
        unit_hydrograph_nrcs[1, 1] = qp
        unit_hydrograph_nrcs[2, 1] = 0

        # vector to store unit hydrograph
        nrcs_unit_hydrograph = np.zeros([len(hydrographs_intervals)])

        # calculate hydrograph
        section = 0
        for j in range(0, len(hydrographs_intervals)):
            time = hydrographs_intervals[j] / 60
            if time <= tp:
                nrcs_unit_hydrograph[j] = np.interp(time, unit_hydrograph_nrcs[:2, 0], unit_hydrograph_nrcs[:2, 1])
            else:
                nrcs_unit_hydrograph[j] = np.interp(time, unit_hydrograph_nrcs[1:, 0], unit_hydrograph_nrcs[1:, 1])

        # update hydrographs dataframe
        nrcs_unit_hydrographs_df.loc[intervals_index, i] = nrcs_unit_hydrograph

        # plt.plot(hydrographs_intervals, nrcs_unit_hydrograph)
        # plt.show()

        print(str(counter) + ' of ' + str(len(cols)))
        counter += 1
    # ==================================================================================================================
    # -- Export unit hydrographs
    # ==================================================================================================================
    book = load_workbook(max_hydrology_file)
    writer = pd.ExcelWriter(max_hydrology_file, engine='openpyxl')
    writer.book = book

    hydrograph_time_intervals_df.to_excel(writer, sheet_name='9. Unit Hydrographs Intervals')
    w_h_unit_hydrographs_df.to_excel(writer, sheet_name='10. W & H Unit Hydrographs')
    snyder_unit_hydrographs_df.to_excel(writer, sheet_name='11. Snyder Unit Hydrographs')
    nrcs_unit_hydrographs_df.to_excel(writer, sheet_name='12. Snyder Unit Hydrographs')

    writer.save()
    writer.close()


def flow_hydrographs(analysis_points, max_hydrology_file, exclusion_file=None):
    """
    Estimate unit Hydrographs.

    Parameters
    ----------
    analysis_points: str
        Analysis points vector
    max_hydrology_file: str
        Excel file containing hydrology analysis of maximum streamflows [must have morphometry and CN]
    exclusion_file: str
        Excel file containing list of excluded points from hydrology analysis

    Returns
    -------
    Williams & Hann, Snyder and NRCS hydrographs and maximum streamflows

    """
    # ==================================================================================================================
    # -- Main imports and preparations
    # ==================================================================================================================
    import sqlite3
    import numpy as np
    import pandas as pd
    import grass.script as grass
    from openpyxl import load_workbook
    import grass.script.array as garray

    # load attribute table as dataframe
    sql_path = grass.read_command('db.databases', driver='sqlite').replace('\n', '')
    con = sqlite3.connect(sql_path)
    sql_stat = 'SELECT * FROM %s ' % analysis_points
    points_df = pd.read_sql_query(sql_stat, con, index_col='cat')

    # excluded points from analysis
    if exclusion_file:
        exclusion_list = pd.read_excel(exclusion_file, header=0)['cat']
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values if i not in exclusion_list.values]
    else:
        cols = ['stream_' + str(int(i)) for i in points_df['value'].values]

    # ==================================================================================================================
    # -- Base info and new Dataframes
    # ==================================================================================================================
    # base info dataframes
    return_periods = [2, 5, 10, 25, 50, 100, 200, 500]
    effective_cumm_depth_distribution_df = pd.read_excel(max_hydrology_file, sheet_name='8F. P Efect Depth Dist',
                                                         header=0, index_col=0)
    hydrograph_time_intervals_df = pd.read_excel(max_hydrology_file, sheet_name='9. Unit Hydrographs Intervals',
                                                 header=0, index_col=0)
    w_h_unit_hydrographs_df = pd.read_excel(max_hydrology_file, sheet_name='10. W & H Unit Hydrographs',
                                            header=0, index_col=0)
    snyder_unit_hydrographs_df = pd.read_excel(max_hydrology_file, sheet_name='11. Snyder Unit Hydrographs',
                                               header=0, index_col=0)
    nrcs_unit_hydrographs_df = pd.read_excel(max_hydrology_file, sheet_name='12. Snyder Unit Hydrographs',
                                             header=0, index_col=0)

    # dataframes to store design storm time distribution
    intervals_index = ['Interval %s' % i for i in range(1, len(w_h_unit_hydrographs_df) + 1)]
    flow_intervals = []
    for i in return_periods:
        for k in intervals_index:
            flow_intervals.append(k + ' - Tr' + str(i))

    # new dataframes
    w_h_hydrographs_df = pd.DataFrame(data=np.zeros([len(return_periods) * len(w_h_unit_hydrographs_df), len(cols)]),
                                      columns=cols, index=flow_intervals)
    snyder_hydrographs_df = pd.DataFrame(data=np.zeros([len(return_periods) * len(w_h_unit_hydrographs_df), len(cols)]),
                                         columns=cols, index=flow_intervals)
    nrcs_hydrographs_df = pd.DataFrame(data=np.zeros([len(return_periods) * len(w_h_unit_hydrographs_df), len(cols)]),
                                       columns=cols, index=flow_intervals)
    extreme_flows_index = []
    for i in return_periods:
        for j in ['W & Hann', 'Snyder', 'NRCS', 'Selected']:
            extreme_flows_index.append(j + ' - Tr' + str(i))
    extreme_flows_df = pd.DataFrame(data=np.zeros([4 * len(return_periods), len(cols)]), columns=cols,
                                    index=extreme_flows_index)

    # ==================================================================================================================
    # -- Calculate Flow Hydrographs [W & H, Snyder, NRCS]
    # ==================================================================================================================
    tr_index_extended = effective_cumm_depth_distribution_df.index.values
    tr_index = [int(i.split('Tr')[1]) for i in tr_index_extended]
    tr_index = np.array(tr_index)

    counter = 1
    for i in cols:
        for j in return_periods:
            tr_index_event = effective_cumm_depth_distribution_df[i].values[np.where(tr_index == j)]
            hydrgraphs_n = len(w_h_unit_hydrographs_df)
            event_matrix = np.zeros([hydrgraphs_n, hydrgraphs_n])
            c = 0
            for n in range(0, hydrgraphs_n):
                if c + len(tr_index_event) < len(event_matrix):
                    event_matrix[c:c + len(tr_index_event), n] = tr_index_event[0:len(tr_index_event)]
                else:
                    event_matrix[c:len(event_matrix), n] = tr_index_event[0:len(event_matrix) - c]
                c += 1

            w_h_unit_hydrograph = w_h_unit_hydrographs_df[i].values
            w_h_flow_hydrograph = np.matmul(event_matrix, w_h_unit_hydrograph)
            snyder_unit_hydrograph = snyder_unit_hydrographs_df[i].values
            snyder_flow_hydrograph = np.matmul(event_matrix, snyder_unit_hydrograph)
            nrcs_unit_hydrograph = nrcs_unit_hydrographs_df[i].values
            nrcs_flow_hydrograph = np.matmul(event_matrix, nrcs_unit_hydrograph)

            intervals_index = ['Interval %s' % t for t in range(1, len(w_h_unit_hydrographs_df) + 1)]
            flow_intervals = []
            for m in intervals_index:
                flow_intervals.append(m + ' - Tr' + str(j))

            w_h_hydrographs_df.loc[flow_intervals, i] = w_h_flow_hydrograph
            snyder_hydrographs_df.loc[flow_intervals, i] = snyder_flow_hydrograph
            nrcs_hydrographs_df.loc[flow_intervals, i] = nrcs_flow_hydrograph

            w_h_max = w_h_flow_hydrograph.max()
            snyder_max = snyder_flow_hydrograph.max()
            nrcs_max = nrcs_flow_hydrograph.max()

            extreme_flows_df.loc['W & Hann - Tr' + str(j), i] = w_h_max
            extreme_flows_df.loc['Snyder - Tr' + str(j), i] = snyder_max
            extreme_flows_df.loc['NRCS - Tr' + str(j), i] = nrcs_max

            max_vector = [w_h_max, snyder_max, nrcs_max]
            stde = np.std(max_vector)
            median = np.median(max_vector)
            mean = np.mean(max_vector)

            if abs(median - mean) / median <= 0.10:
                max_flow = mean
            else:
                c = 0
                max_flow = 0
                for k in max_vector:
                    if k <= median + stde and k >= median - stde:
                        max_flow += k
                        c += 1
                if c >= 2:
                    max_flow = max_flow / c
                else:
                    max_flow = np.max(max_vector)

            extreme_flows_df.loc['Selected - Tr' + str(j), i] = max_flow

        print(str(counter) + ' of ' + str(len(cols)))
        counter += 1

    # ==================================================================================================================
    # -- Export unit hydrographs
    # ==================================================================================================================
    book = load_workbook(max_hydrology_file)
    writer = pd.ExcelWriter(max_hydrology_file, engine='openpyxl')
    writer.book = book

    w_h_hydrographs_df.to_excel(writer, sheet_name='13. W & Hann Hydrographs')
    snyder_hydrographs_df.to_excel(writer, sheet_name='14. Snyder Hydrographs')
    nrcs_hydrographs_df.to_excel(writer, sheet_name='15. NRCS Hydrographs')
    extreme_flows_df.to_excel(writer, sheet_name='16. Extreme Flows')

    writer.save()
    writer.close()


def idf_wilches(timeseries, model='B'):
    """IDF curves by Wilches methodology.

    Parameters
    ----------
    timeseries : [type]
        [description]
    scaling_factor : float, optional
        [description], by default 0.85

    Returns
    ----------
    pd.Dataframe
        Intensity, Duration, Frequency curves

    """
    # ==================================================================================================================
    # -- Main imports and preparations
    # ==================================================================================================================
    import numpy as np
    import pandas as pd
    from math import exp
    from scipy.stats import norm

    # ==================================================================================================================
    # -- IDF curves Wilches methodology
    # ==================================================================================================================
    # estimate frequencies
    return_periods = [2.33, 5, 10, 25, 50, 100, 200, 500]
    probability = [1 / i for i in return_periods]
    inverse_probability = [1 - i for i in probability]

    inverse_gaussian = [norm.ppf(i) for i in inverse_probability]

    # duration
    reference_duration = 1440
    if model == 'A':
        first_duration_interval = np.arange(5, 45, 5)
        second_duration_interval = np.arange(45, 105, 5)
        third_duration_interval = np.arange(105, 1445, 5)
        duration_interval = np.concatenate([first_duration_interval, second_duration_interval, third_duration_interval])
    elif model == 'B':
        first_duration_interval = np.arange(5, 105, 5)
        second_duration_interval = np.arange(105, 1445, 5)
        duration_interval = np.concatenate([first_duration_interval, second_duration_interval])
    else:
        print('!not a valid model')

    # dataframe to store IDF curves
    idf_df = pd.DataFrame(index=duration_interval, columns=return_periods)

    try:
        # basic statistics
        mean = timeseries.mean()
        standar_deviation = timeseries.std()
        cv = standar_deviation / mean

        # estimate intensities
        for i in range(0, len(return_periods)):
            if model == 'A':
                for j in third_duration_interval:
                    z = inverse_gaussian[i]
                    intensity = mean * (exp(z * np.sqrt(np.log(1 + cv ** 2))) / np.sqrt(1 + cv ** 2)) * \
                        (j / reference_duration) ** -0.85
                    idf_df.loc[j][idf_df.columns.values[i]] = intensity

                for j in second_duration_interval:
                    z = inverse_gaussian[i]
                    intensity = mean * (exp(z * np.sqrt(np.log(1 + cv ** 2))) / np.sqrt(1 + cv ** 2)) * \
                        (j / reference_duration) ** -0.66
                    idf_df.loc[j][idf_df.columns.values[i]] = intensity

                reference_intensity = idf_df.loc[60][idf_df.columns.values[i]]

                for j in first_duration_interval:
                    intensity = reference_intensity * (32.4 / j ** 0.75 - 30 / j)
                    idf_df.loc[j][idf_df.columns.values[i]] = intensity

            elif model == 'B':
                for j in second_duration_interval:
                    z = inverse_gaussian[i]
                    intensity = mean * (exp(z * np.sqrt(np.log(1 + cv ** 2))) / np.sqrt(1 + cv ** 2)) * \
                        (j / reference_duration) ** -0.85
                    idf_df.loc[j][idf_df.columns.values[i]] = intensity

                reference_intensity = idf_df.loc[105][idf_df.columns.values[i]]

                for j in first_duration_interval:
                    intensity = reference_intensity * (46.2 / j ** 0.75 - 43.05 / j)
                    idf_df.loc[j][idf_df.columns.values[i]] = intensity

        return idf_df

    except AttributeError:
        return None


def idf_pulgarin(timeseries):
    """IDF curves by Pulgarin methodology [Andean Region].

    Parameters
    ----------
    timeseries : [type]
        [description]
    scaling_factor : float, optional
        [description], by default 0.85

    Returns
    ----------
    pd.Dataframe
        Intensity, Duration, Frequency curves

    """
    # ==================================================================================================================
    # -- Main imports and preparations
    # ==================================================================================================================
    import numpy as np
    import pandas as pd
    from math import exp
    from scipy.stats import norm

    # ==================================================================================================================
    # -- IDF curves Pulgarin methodology
    # ==================================================================================================================
    # estimate frequencies
    return_periods = [2.33, 5, 10, 25, 50, 100, 200, 500]
    probability = [1 / i for i in return_periods]
    inverse_probability = [1 - i for i in probability]

    inverse_gaussian = [norm.ppf(i) for i in inverse_probability]

    # duration
    reference_duration = 60
    first_duration_interval = np.arange(5, 60, 2.5)
    second_duration_interval = np.arange(60, 1445, 5)
    duration_interval = np.concatenate([first_duration_interval, second_duration_interval])

    # dataframe to store IDF curves
    idf_df = pd.DataFrame(index=duration_interval, columns=return_periods)

    try:
        # basic statistics
        mean = timeseries.mean()
        standar_deviation = timeseries.std()
        cv = standar_deviation / mean

        # estimate intensities
        for i in range(0, len(return_periods)):
            for j in second_duration_interval:
                z = inverse_gaussian[i]
                z = -np.log((-np.log( 1 - 1 / i)))
                intensity = mean * (exp(z * np.sqrt(np.log(1 + cv ** 2))) / np.sqrt(1 + cv ** 2)) * \
                    (j / reference_duration) ** -0.85
                idf_df.loc[j][idf_df.columns.values[i]] = intensity

            for j in second_duration_interval:
                z = inverse_gaussian[i]
                intensity = mean * (exp(z * np.sqrt(np.log(1 + cv ** 2))) / np.sqrt(1 + cv ** 2)) * \
                    (j / reference_duration) ** -0.66
                idf_df.loc[j][idf_df.columns.values[i]] = intensity

            reference_intensity = idf_df.loc[60][idf_df.columns.values[i]]

            for j in first_duration_interval:
                intensity = reference_intensity * (32.4 / j ** 0.75 - 30 / j)
                idf_df.loc[j][idf_df.columns.values[i]] = intensity

        return idf_df

    except AttributeError:
        return None